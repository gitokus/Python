import csv

from openpyxl import load_workbook


def save_list_to_csv(a_list_of_lists, a_file, endline_type='\n'):
    with open(a_file, 'w') as csvfile:
        fcsvwriter = csv.writer(csvfile, dialect='excel',
                                lineterminator=endline_type)
        for l in a_list_of_lists:
            if isinstance(l, list):
                fcsvwriter.writerow(l)
            else:
                fcsvwriter.writerow([l])


def compare_sheets(xl_sheet_1, xl_sheet_2):
    # picking up row 2 (with headers)
    header_row_1 = [c.value for c in xl_sheet_1['2']]
    header_row_2 = [c.value for c in xl_sheet_2['2']]

    # header which needs to be the same
    header_key = header_row_1[1:26]

    # index of Signal Name (aka column C)
    signal_name_idx = header_row_1.index('Signal Name')

    # list with signal names from older version
    signal_names_1 = [c.value for c in xl_sheet_1['C']]
    signal_names_2 = [c.value for c in xl_sheet_2['C']]
    removed_signals = list(set(signal_names_1) - set(signal_names_2))

    status_vector = [['Signal Name', 'Changes', 'Strikethrough status']]
    signal_presence = 'init'

    for row in xl_sheet_2.iter_rows():
        signal_name = row[signal_name_idx].value
        if signal_name and signal_name != \
                'Signal Name':
            try:
                row_idx_in_1 = signal_names_1.index(signal_name) + 1
            except ValueError:
                # new signal
                if row[signal_name_idx].font.strike:
                    signal_presence = 'None/Strikethrough'
                else:
                    signal_presence = 'None/Clear'
                status_vector.append([signal_name, 'New signal', signal_presence])
                continue

            row_from_1 = xl_sheet_1[row_idx_in_1]
            strike_status_2 = row[signal_name_idx].font.strike
            strike_status_1 = xl_sheet_1.cell(row=row_idx_in_1,
                                              column=signal_name_idx).font.strike
            if strike_status_2 and strike_status_1:
                signal_presence = 'Strikethrough/Strikethrough'
            elif not strike_status_2 and strike_status_1:
                signal_presence = 'Strikethrough/Clear'
            elif strike_status_2 and not strike_status_1:
                signal_presence = 'Clear/Strikethrough'
            elif not strike_status_2 and not strike_status_1:
                signal_presence = 'Clear/Clear'

            values_from_row2 = [c.value for c in row[1:26]]
            values_from_row1 = [c.value for c in row_from_1[1:26]]
            if values_from_row1 == values_from_row2:
                # the same in both files
                status_vector.append(
                    [signal_name, 'No changes', signal_presence])
            else:
                # different
                comparison_vector = [x != y for (x, y) in
                                     zip(values_from_row1, values_from_row2)]
                change_summary = ''
                for sig_idx in range(0, len(comparison_vector)):
                    if comparison_vector[sig_idx]:
                        if not values_from_row1[sig_idx]:
                            v1 = 'Empty'
                        else:
                            v1 = str(values_from_row1[sig_idx])
                        if not values_from_row2[sig_idx]:
                            v2 = 'Empty'
                        else:
                            v2 = str(values_from_row2[sig_idx])

                        change_summary += header_key[sig_idx] + ': ' + \
                                          v1 + ' to ' + \
                                          v2 + '; '

                status_vector.append(
                    [signal_name, change_summary, signal_presence])

    for row in removed_signals:
        status_vector.append([row, 'Signal completely removed', '?/None'])

    return status_vector


"""
Tested in Python 3.6.4
This version is comparing v004 with v005. Modifications might be needed in case of newer versions.
"""

if __name__ == '__main__':
    # input_excel_path_1 = sys.argv[1]
    # input_excel_path_2 = sys.argv[2]

    input_excel_path_1 = r"C:\Users\qzmwdh\Documents\_DAT2\LROS_reformatting\DAT2p0_AptivToFord_InterfaceSpec.xlsx"
    input_excel_path_2 = r"C:\Users\qzmwdh\Documents\_DAT2\LROS_reformatting\Stage_E_v005\DAT2p0_AptivToFord_InterfaceSpec_v005.xlsx"

    # loading excel file
    xl_workbook_1 = load_workbook(input_excel_path_1)
    xl_workbook_2 = load_workbook(input_excel_path_2)

    # selecting proper tab
    aptiv_to_ford_old = xl_workbook_1['Aptiv_Ford_Interface']
    aptiv_to_ford_new = xl_workbook_2['AptivToFord_Interface']

    # selecting proper tab
    ford_to_aptiv_old = xl_workbook_1['Ford_Aptiv_Interface']
    ford_to_aptiv_new = xl_workbook_2['FordToAptiv_Interface']

    status_vector_aptiv_to_ford = compare_sheets(aptiv_to_ford_old,
                                                 aptiv_to_ford_new)
    status_vector_ford_to_aptiv = compare_sheets(ford_to_aptiv_old,
                                                 ford_to_aptiv_new)

    save_list_to_csv(status_vector_aptiv_to_ford,
                     r'C:\Users\qzmwdh\Documents\_DAT2\LROS_reformatting\Stage_E_v005\DAT2p0_AptivToFord_InterfaceSpec.csv')
    save_list_to_csv(status_vector_ford_to_aptiv,
                     r'C:\Users\qzmwdh\Documents\_DAT2\LROS_reformatting\Stage_E_v005\DAT2p0_FordToAptiv_InterfaceSpec.csv')
